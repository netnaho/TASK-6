import csv
import io
import json
import logging
import uuid

from openpyxl import Workbook, load_workbook

from app.core.config import get_settings
from app.core.constants import DeclarationState, DeliveryFileType, Role
from app.core.exceptions import NotFoundError, ValidationError
from app.models.declaration import DeclarationPackage
from app.models.import_export import ExportJob, FieldMapping, ImportJob, MaskingPolicy
from app.models.plan import NutritionPlan
from app.models.profile import ParticipantProfile
from app.repositories.delivery_repository import DeliveryRepository
from app.models.user import User
from app.repositories.import_export_repository import ImportExportRepository
from app.services.audit_service import AuditService
from app.services.declaration_service import DeclarationService
from app.services.delivery_service import DeliveryService
from app.services.profile_service import ProfileService
from app.schemas.profiles import ProfileUpsertRequest, SensitiveProfilePayload
from app.storage.file_manager import FileManager
from app.utils.datetime import utc_now


logger = logging.getLogger(__name__)


class ImportExportService:
    def __init__(self, db):
        self.db = db
        self.repo = ImportExportRepository(db)
        self.delivery_repo = DeliveryRepository(db)
        self.files = FileManager()
        self.audit = AuditService(db)
        self.settings = get_settings()
        self.declarations = DeclarationService(db)
        self.profiles = ProfileService(db)

    def _create_admin_file(self, *, filename: str, folder: str, content: bytes, mime_type: str, file_type: str, created_by: str):
        storage_path, size_bytes, checksum = self.files.write_bytes(folder, filename, content)
        file = DeliveryService(self.db).create_standalone_file(
            created_by=created_by,
            display_name=filename,
            file_type=file_type,
            storage_path=storage_path,
            mime_type=mime_type,
            checksum_sha256=checksum,
            size_bytes=size_bytes,
            allowed_roles=[Role.ADMINISTRATOR],
            version_label="admin-job",
        )
        return file, storage_path, checksum

    def create_mapping(self, user, payload):
        mapping = FieldMapping(**payload.model_dump(), created_by=user.username)
        self.db.add(mapping)
        self.db.commit()
        self.db.refresh(mapping)
        return mapping

    def create_masking_policy(self, user, payload):
        policy = MaskingPolicy(**payload.model_dump(), created_by=user.username)
        self.db.add(policy)
        self.db.commit()
        self.db.refresh(policy)
        return policy

    def list_mappings(self):
        return self.repo.list_mappings()

    def list_masking_policies(self):
        return self.repo.list_masking_policies()

    async def import_file(self, user, upload, format_name: str, mapping_id=None, scope_type: str = "declarations"):
        self._import_actor = user
        content = await upload.read()
        source_file, storage_path, checksum = self._create_admin_file(
            filename=upload.filename,
            folder="imports",
            content=content,
            mime_type=upload.content_type or "application/octet-stream",
            file_type=DeliveryFileType.IMPORT_SOURCE,
            created_by=user.username,
        )
        rows = self._read_rows(content, format_name)
        rows = self.apply_import_mapping(rows, mapping_id)
        errors: list[dict] = []
        success_count = 0
        failure_count = 0

        for index, row in enumerate(rows, start=1):
            try:
                with self.db.begin_nested():
                    if scope_type == "declarations":
                        self._import_declaration_row(user, row)
                    elif scope_type == "profiles":
                        self._import_profile_row(user, row)
                    else:
                        raise ValidationError("Unsupported import scope.")
                success_count += 1
            except Exception as exc:
                failure_count += 1
                errors.append({"row": index, "error": str(exc), "row_data": row})

        error_report_path = None
        if errors:
            error_report_path, _, _ = self.files.write_bytes("imports", f"import-errors-{uuid.uuid4().hex}.json", json.dumps(errors, default=str).encode())

        job = ImportJob(
            created_by=user.id,
            format=format_name,
            source_file_id=source_file.id,
            mapping_id=mapping_id,
            status="completed",
            row_count=len(rows),
            success_count=success_count,
            failure_count=failure_count,
            checksum_sha256=checksum,
            started_at=utc_now(),
            completed_at=utc_now(),
            error_report_path=error_report_path or storage_path,
        )
        self.db.add(job)
        self.db.flush()
        self.audit.log(actor_user_id=user.id, action_type="data_import", entity_type="import_job", entity_id=str(job.id), metadata={"scope_type": scope_type, "row_count": len(rows), "success_count": success_count, "failure_count": failure_count})
        self.audit.log(actor_user_id=user.id, action_type="import_completed", entity_type="import_job", entity_id=str(job.id), metadata={"scope_type": scope_type, "row_count": len(rows), "success_count": success_count, "failure_count": failure_count})
        logger.info(
            "Data imported",
            extra={
                "job_id": str(job.id),
                "format": format_name,
                "scope_type": scope_type,
                "record_count": len(rows),
                "success_count": success_count,
                "failure_count": failure_count,
            },
        )
        self.db.commit()
        self.db.refresh(job)
        return job

    def export_rows(self, user, format_name: str, scope_type: str, rows: list[dict], masking_policy_id=None, mapping_id=None):
        masked_rows = self.apply_masking(rows, masking_policy_id)
        content, filename = self._write_rows(masked_rows, format_name, scope_type, mapping_id)
        output_file, _, checksum = self._create_admin_file(
            filename=filename,
            folder="exports",
            content=content,
            mime_type="text/csv" if format_name == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            file_type=DeliveryFileType.EXPORT_ARTIFACT,
            created_by=user.username,
        )
        job = ExportJob(
            created_by=user.id,
            format=format_name,
            scope_type=scope_type,
            masking_policy_id=masking_policy_id,
            status="completed",
            row_count=len(masked_rows),
            checksum_sha256=checksum,
            output_file_id=output_file.id,
            started_at=utc_now(),
            completed_at=utc_now(),
        )
        self.db.add(job)
        self.db.flush()
        self.audit.log(actor_user_id=user.id, action_type="data_export", entity_type="export_job", entity_id=str(job.id), metadata={"scope_type": scope_type, "row_count": len(masked_rows), "format": format_name})
        logger.info(
            "Data exported",
            extra={"job_id": str(job.id), "format": format_name, "scope_type": scope_type, "record_count": len(masked_rows)},
        )
        self.db.commit()
        self.db.refresh(job)
        return job

    def export_scope(self, user, format_name: str, scope_type: str, masking_policy_id=None, mapping_id=None):
        if scope_type == "declarations":
            packages = self.db.query(DeclarationPackage).order_by(DeclarationPackage.created_at.desc()).all()
            rows = [
                {
                    "package_number": item.package_number,
                    "participant_id": str(item.participant_id),
                    "state": str(item.state),
                    "submitted_at": item.submitted_at.isoformat() if item.submitted_at else None,
                    "review_due_at": item.review_due_at.isoformat() if item.review_due_at else None,
                }
                for item in packages
            ]
            return self.export_rows(user, format_name, scope_type, rows, masking_policy_id, mapping_id)
        if scope_type == "profiles":
            if format_name != "xlsx":
                raise ValidationError("Profile exports are only supported in XLSX format.")
            profiles = self.db.query(ParticipantProfile).order_by(ParticipantProfile.created_at.desc()).all()
            rows = [self._profile_export_row(item) for item in profiles]
            return self.export_rows(user, format_name, scope_type, rows, masking_policy_id, mapping_id)
        raise ValidationError("Unsupported export scope.")

    def get_import_detail(self, job_id):
        job = self.repo.get_import(job_id)
        if not job:
            raise NotFoundError("Import job not found.")
        source_file = self.delivery_repo.get_file(job.source_file_id) if job.source_file_id else None
        preview_rows = self._read_preview_rows(source_file.storage_path, job.format) if source_file else []
        errors: list[dict] = []
        if job.error_report_path and self.files.exists(job.error_report_path):
            try:
                errors = json.loads(self.files.read_bytes(job.error_report_path).decode())
            except (UnicodeDecodeError, json.JSONDecodeError):
                errors = []
        return {"job": job, "source_file": source_file, "errors": errors, "preview_rows": preview_rows}

    def get_export_detail(self, job_id):
        job = self.repo.get_export(job_id)
        if not job:
            raise NotFoundError("Export job not found.")
        output_file = self.delivery_repo.get_file(job.output_file_id) if job.output_file_id else None
        preview_rows = self._read_preview_rows(output_file.storage_path, job.format) if output_file else []
        return {"job": job, "output_file": output_file, "preview_rows": preview_rows}

    def create_import_source_download_link(self, user, job_id):
        job = self.repo.get_import(job_id)
        if not job or not job.source_file_id:
            raise NotFoundError("Import source file not found.")
        return DeliveryService(self.db).create_direct_download_link(user, job.source_file_id, issued_to_user_id=user.id, purpose="import_source_download")

    def create_export_download_link(self, user, job_id):
        job = self.repo.get_export(job_id)
        if not job or not job.output_file_id:
            raise NotFoundError("Export artifact not found.")
        return DeliveryService(self.db).create_direct_download_link(user, job.output_file_id, issued_to_user_id=user.id, purpose="export_download")

    def apply_import_mapping(self, rows: list[dict], mapping_id):
        if not mapping_id:
            return rows
        mapping = self.repo.get_mapping(mapping_id)
        if not mapping:
            raise NotFoundError("Field mapping not found.")
        reverse_mapping = {str(value): key for key, value in mapping.mapping_json.items()}
        transformed: list[dict] = []
        for row in rows:
            transformed.append({reverse_mapping.get(key, key): value for key, value in row.items()})
        return transformed

    def apply_export_mapping(self, rows: list[dict], mapping_id):
        if not mapping_id:
            return rows
        mapping = self.repo.get_mapping(mapping_id)
        if not mapping:
            raise NotFoundError("Field mapping not found.")
        transformed: list[dict] = []
        for row in rows:
            transformed.append({mapping.mapping_json.get(key, key): value for key, value in row.items()})
        return transformed

    def apply_masking(self, rows: list[dict], masking_policy_id):
        if not masking_policy_id:
            return rows
        policy = self.repo.get_masking_policy(masking_policy_id)
        if not policy:
            raise NotFoundError("Masking policy not found.")
        rules = policy.rules_json
        masked_rows: list[dict] = []
        for row in rows:
            masked = dict(row)
            for field in rules.get("mask_fields", []):
                if field in masked:
                    masked[field] = self.settings.export_mask
            masked_rows.append(masked)
        return masked_rows

    def _read_preview_rows(self, storage_path: str, format_name: str, limit: int = 10) -> list[dict]:
        if not self.files.exists(storage_path):
            return []
        rows = self._read_rows(self.files.read_bytes(storage_path), format_name)
        return rows[:limit]

    @staticmethod
    def _coerce_json_value(value, default):
        if value in (None, ""):
            return default
        if isinstance(value, (dict, list)):
            return value
        if isinstance(value, str):
            try:
                return json.loads(value)
            except json.JSONDecodeError:
                return default
        return default

    def _profile_export_row(self, profile: ParticipantProfile) -> dict:
        sensitive = profile.sensitive
        return {
            "user_id": str(profile.user_id),
            "profile_status": profile.profile_status,
            "demographics_json": json.dumps(profile.demographics_json),
            "medical_flags_json": json.dumps(profile.medical_flags_json),
            "activity_json": json.dumps(profile.activity_json),
            "anthropometrics_json": json.dumps(profile.anthropometrics_json),
            "diagnoses_notes": sensitive.get("diagnoses_notes"),
            "medications_detail": sensitive.get("medications_detail"),
            "allergy_detail": sensitive.get("allergy_detail"),
            "pregnancy_detail": sensitive.get("pregnancy_detail"),
            "clinician_notes": sensitive.get("clinician_notes"),
            "sensitive_free_text": sensitive.get("sensitive_free_text"),
        }

    def _read_rows(self, content: bytes, format_name: str) -> list[dict]:
        if format_name == "csv":
            text = content.decode()
            return list(csv.DictReader(io.StringIO(text)))
        if format_name == "xlsx":
            workbook = load_workbook(io.BytesIO(content))
            sheet = workbook.active
            if sheet is None:
                return []
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] for i in range(len(headers))})
            return rows
        raise ValidationError("Unsupported format.")

    def _write_rows(self, rows: list[dict], format_name: str, scope_type: str, mapping_id=None) -> tuple[bytes, str]:
        rows = self.apply_export_mapping(rows, mapping_id)
        if format_name == "csv":
            stream = io.StringIO()
            writer = csv.DictWriter(stream, fieldnames=list(rows[0].keys()) if rows else ["empty"])
            writer.writeheader()
            if rows:
                writer.writerows(rows)
            return stream.getvalue().encode(), f"{scope_type}.csv"
        if format_name == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            if sheet is None:
                raise ValidationError("Unable to create workbook sheet.")
            headers = list(rows[0].keys()) if rows else ["empty"]
            sheet.append(headers)
            for row in rows:
                sheet.append([row.get(header) for header in headers])
            buffer = io.BytesIO()
            workbook.save(buffer)
            return buffer.getvalue(), f"{scope_type}.xlsx"
        raise ValidationError("Unsupported format.")

    def _import_declaration_row(self, actor, row: dict) -> None:
        participant_id = row.get("participant_id")
        if not participant_id:
            raise ValidationError("participant_id is required for declaration import.")
        participant = self.db.query(User).filter(User.id == participant_id).one_or_none()
        if not participant:
            raise ValidationError("Referenced participant does not exist.")
        profile_id = row.get("profile_id")
        plan_id = row.get("plan_id")
        profile = self.db.query(ParticipantProfile).filter(ParticipantProfile.id == profile_id, ParticipantProfile.user_id == participant.id).one_or_none()
        plan = self.db.query(NutritionPlan).filter(NutritionPlan.id == plan_id, NutritionPlan.participant_id == participant.id).one_or_none()
        if not profile:
            raise ValidationError("Imported declaration requires a valid profile_id for the participant.")
        if not plan:
            raise ValidationError("Imported declaration requires a valid plan_id for the participant.")
        requested_state = DeclarationState(str(row.get("state", DeclarationState.DRAFT)))
        package = self.declarations.create(
            participant,
            profile.id,
            plan.id,
            actor_user=actor,
            package_number=row.get("package_number") or f"IMPORTED-{uuid.uuid4().hex[:10].upper()}",
            commit=False,
        )
        if requested_state == DeclarationState.DRAFT:
            return
        if requested_state == DeclarationState.SUBMITTED:
            self.declarations.transition(actor, package.id, DeclarationState.SUBMITTED, reason_code="imported_record", reason_text="Imported declaration record.", commit=False)
            return
        if requested_state == DeclarationState.WITHDRAWN:
            self.declarations.transition(actor, package.id, DeclarationState.SUBMITTED, reason_code="imported_record", reason_text="Imported declaration record.", commit=False)
            self.declarations.transition(actor, package.id, DeclarationState.WITHDRAWN, reason_code="imported_record", reason_text="Imported declaration record.", commit=False)
            return
        if requested_state == DeclarationState.CORRECTED:
            self.declarations.transition(actor, package.id, DeclarationState.SUBMITTED, reason_code="imported_record", reason_text="Imported declaration record.", commit=False)
            self.declarations.transition(actor, package.id, DeclarationState.CORRECTED, reason_code="imported_record", reason_text="Imported declaration record.", commit=False)
            return
        if requested_state == DeclarationState.VOIDED:
            self.declarations.transition(actor, package.id, DeclarationState.VOIDED, reason_code="imported_record", reason_text="Imported declaration record.", commit=False)
            return
        raise ValidationError("Unsupported declaration import state.")

    def _import_profile_row(self, actor, row: dict) -> None:
        user_id = row.get("user_id")
        if not user_id:
            raise ValidationError("user_id is required for profile import.")
        target_user = self.db.query(User).filter(User.id == user_id).one_or_none()
        if not target_user:
            raise ValidationError("Referenced user does not exist.")
        payload = ProfileUpsertRequest(
            profile_status=row.get("profile_status", "in_progress"),
            demographics_json=self._coerce_json_value(row.get("demographics_json"), {}),
            medical_flags_json=self._coerce_json_value(row.get("medical_flags_json"), {}),
            activity_json=self._coerce_json_value(row.get("activity_json"), {}),
            anthropometrics_json=self._coerce_json_value(row.get("anthropometrics_json"), {}),
            sensitive=SensitiveProfilePayload(
                diagnoses_notes=row.get("diagnoses_notes"),
                medications_detail=row.get("medications_detail"),
                allergy_detail=row.get("allergy_detail"),
                pregnancy_detail=row.get("pregnancy_detail"),
                clinician_notes=row.get("clinician_notes"),
                sensitive_free_text=row.get("sensitive_free_text"),
            ),
        )
        self.profiles.import_for_user(actor, target_user, payload)
